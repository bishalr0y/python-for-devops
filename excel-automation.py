import openpyxl
inv_file = openpyxl.load_workbook("inventory.xlsx")

print(inv_file.sheetnames)
product_list = inv_file['Sheet1']
#print(product_list)

# List each company with respective product count
products_per_supplier = {}
for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    
    if supplier_name in products_per_supplier:
        products_per_supplier[supplier_name] = products_per_supplier[supplier_name] + 1
    else:
        products_per_supplier[supplier_name] = 1

print(products_per_supplier)


# List products with inventory less than 10
product_inv_lt10 = {}
for product_row in range(2, product_list.max_row + 1):
    product_count = product_list.cell(product_row, 2).value
    
    if product_count < 10:
        product_inv_lt10[product_list.cell(product_row, 1).value] = product_count
        
print(product_inv_lt10)



# List each company with respective total inventory value
total_value_per_supplier = {}
for product_row in range(2, product_list.max_row + 1):
    supplier_name = product_list.cell(product_row, 4).value
    
    if supplier_name in total_value_per_supplier:
        total_value_per_supplier[supplier_name] = total_value_per_supplier[supplier_name] + product_list.cell(product_row, 2).value * product_list.cell(product_row, 3).value
    else:
        total_value_per_supplier[supplier_name] = product_list.cell(product_row, 2).value * product_list.cell(product_row, 3).value


print(total_value_per_supplier)

# Write to spreadsheet: Calculate and write inventory value for each product into spreadsheet


for product_row in range(2, product_list.max_row + 1):
    inv_price = product_list.cell(product_row, 5)
    
    inv_price.value = product_list.cell(product_row, 2).value * product_list.cell(product_row, 3).value

inv_file.save('inventory.xlsx')



